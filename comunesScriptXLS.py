import os
import logging
import cx_Oracle

from openpyxl import load_workbook
from cert_prof_xls_import import CertProf


logging.info('Start scriptXLS.py')

colores_no_validos = (-0.1499984740745262, -0.249977111117893)
filas_no_validas = []


def fila_valida(row, fileName):

    if hasattr(row[1].fill, 'fgColor'):
        if row[1].fill.fgColor.tint in colores_no_validos:
            # logger.info("False")
            # logger.info("RGB: " + str(row[1].fill.fgColor.rgb))
            # logger.info("Indexed: " + str(row[1].fill.fgColor.indexed))
            # logger.info("auto: " + str(row[1].fill.fgColor.auto))
            # logger.info("theme: " + str(row[1].fill.fgColor.theme))
            # logger.info("tint: " + str(row[1].fill.fgColor.tint))
            # logger.info("type: " + str(row[1].fill.fgColor.type))
            response = False
        else:
            # logger.info("True")
            # logger.info("RGB: " + str(row[1].fill.fgColor.rgb))
            # logger.info("Indexed: " + str(row[1].fill.fgColor.indexed))
            # logger.info("auto: " + str(row[1].fill.fgColor.auto))
            # logger.info("theme: " + str(row[1].fill.fgColor.theme))
            # logger.info("tint: " + str(row[1].fill.fgColor.tint))
            # logger.info("type: " + str(row[1].fill.fgColor.type))
            response = True
    else:
        # logger.info("No tiene fgColor")
        response = True

    return response

def fila_vacia(row):
    if all(c.value is None for c in row):
        res = True
    else:
        res = False

    return res

def borra_registros_anteriores(cursor, con):
    sql1 = "DELETE FROM SI_CEPR_DOCUM WHERE AUD_USR_INS = 'elivecro'"
    sql2 = "DELETE FROM SI_CEPR_UNI_COM WHERE AUD_USR_INS = 'elivecro'"
    sql3 = "DELETE FROM SI_CEPR_ESP_FOR_MOD WHERE AUD_USR_INS = 'elivecro'"
    sql4 = "DELETE FROM SI_SOL_CERT_PROF WHERE AUD_USR_INS = 'elivecro'"
    sql5 = "DELETE FROM SI_PER WHERE AUD_USR_INS = 'elivecro'"

    try:
        cursor.execute(sql1)
        cursor.execute(sql2)
        cursor.execute(sql3)
        cursor.execute(sql4)
        cursor.execute(sql5)
        con.commit()
    except cx_Oracle.IntegrityError as e:
        pass
        # error = Error()
        # error.description = " No se han podido eliminar los registros anteriores " + str(
        #     e.args[0].code) + " : " + e.args[0].message
        # error.function = "borraRegistrosAnteriores"
        # error.log_error()
    except cx_Oracle.DatabaseError as e:
        pass
        # error = Error()
        # error.description = " No se han podido eliminar los registros anteriores " + str(
        #     e.args[0].code) + " : " + e.args[0].message
        # error.function = "borraRegistrosAnteriores"
        # error.log_error()


XLS_DIR = "./xls"
START_INDEX = 6  # Número en el que empizan los datos dentro de las hojas excel para el comun

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

INFO_LOG = 'log/INFO.log'
ERROR_LOG = 'log/ERROR.log'

formatter = logging.Formatter('%(levelname)s - %(message)s')

#to log debug messages
debug_log = logging.FileHandler(os.path.join(INFO_LOG))
debug_log.setLevel(logging.INFO)
debug_log.setFormatter(formatter)

#to log errors messages
error_log = logging.FileHandler(os.path.join(ERROR_LOG))
error_log.setLevel(logging.ERROR)
error_log.setFormatter(formatter)

logger.addHandler(debug_log)
logger.addHandler(error_log)

# Connect to the local database.
# connection = pymysql.connect(host='localhost',
#                              user='xls',
#                              password='password',
#                              db='prueba',
#                              charset='utf8mb4',
#                              cursorclass=pymysql.cursors.DictCursor)
#
# cursor = connection.cursor()
#Solo la primera vez
# cursor.execute(BBDD.create_table)
# cursor.execute(BBDD.clean_db)
# connection.commit()

#Conexion base de datos Oracle
conOracle = cx_Oracle.connect('sgf2/sgf2@des')
cursorOracle = conOracle.cursor()

borra_registros_anteriores(cursorOracle, conOracle)

fileList = []
for ruta, dirs, archivos in os.walk(XLS_DIR, topdown=True):
    for a in archivos:
        if "~" not in a:
            fileList.append(ruta.replace("\\", "/") + "/" + a)

# Filtramos para que solo coja los xlsx, tb se eliminan los xlsx abiertos
fileList = filter(lambda file: not file.startswith("~") and (file.endswith("xlsx")), fileList)

for fileName in fileList:
    logger.info('\n' + fileName + ' va a ser parseado ')
    certProf = CertProf()

    try:
        wb = load_workbook(fileName, read_only=True)
        certProf.d_pestana = wb.sheetnames[0]
        ws = wb[certProf.d_pestana]
        rows = ws.rows

        logger.info(fileName + "Numero de filas " + str(ws.max_row))
        idx = 0
        contadorVacias = 0
        countOracle = 0  # Inserciones hechas en des y sqlPrint generadas

        for idx, row in enumerate(rows):
            if idx >= START_INDEX:
                certProf.n_fila = idx + 1

                if all(c.value is None for c in row):
                    logger.info(fileName + "la fila " + str(certProf.n_fila) + " esta vacía " + str(contadorVacias))
                    contadorVacias += 1
                    if contadorVacias >= 3:
                        break
                elif not fila_valida(row, fileName): #  TODO: Filename solo para prueba
                    logger.info(fileName + " : La fila " + str(certProf.n_fila) + " esta marcada como no tratar y no se guardará ")
                    filas_no_validas.append(certProf.n_fila)
                else:
                    contadorVacias = 0
                    logger.info(fileName + " : Procesando fila " + str(certProf.n_fila))

                    certProf.d_excel = fileName

                    certProf.f_present = row[0].value  # date,
                    certProf.set_dni(str(row[1].value).zfill(9))  # varchar2(9),
                    certProf.d_apel1 = row[2].value.strip() if row[2].value else ""   # varchar2(40),
                    certProf.d_apel2 = row[3].value.strip() if row[3].value else ""  # varchar2(40),
                    certProf.d_nombre = row[4].value.strip() if row[4].value else ""  # varchar2(40),
                    certProf.set_sexo(str(row[5].value)) #char(1) Solo puede figurar H o M OK
                    certProf.f_nacim = row[6].value  # date,
                    certProf.d_direccion = str(row[7].value.strip()).replace("'","") if row[7].value else ""  # varchar2(100),
                    certProf.set_cp(str(row[8].value).zfill(5))  # varchar2(5) Comprobar 5 dígitos
                    certProf.set_localidad(str(row[9].value), cursorOracle)  # varchar2(50) No Vacío OK
                    certProf.set_provincia(str(row[10].value.strip() if row[10].value else ""))  # varchar2(40), No Vacío ENUM OK
                    certProf.set_via_expedicion(row[11].value, 1)  # number,Solo 2,6,9 o vacía (Si vacía colocar el número asignado a sin acceso en BBDD) OK
                    certProf.set_d_mods(row[19].value)  # varchar2(40), Solo puede figurar COMPLETO, PARCIAL, UF, INCOMPLETO O VACÍO Esto es modulos formativos columna T  es C_TIPO_SOLICITUD
                    certProf.set_itinerario()
                    certProf.set_c_certificado(row[12].value)
                    certProf.d_real_decreto = row[14].value  # varchar2(100),
                    certProf.d_certificado = row[15].value  # varchar2(100),
                    certProf.f_expedicion = row[17].value  # date,
                    certProf.d_fcs = row[18].value  # varchar2(100),
                    certProf.d_doc_present = row[20].value  # varchar2(40),
                    certProf.d_observaciones = str(row[22].value).replace("--", " ").replace("'", " ")  # varchar2(512),
                    certProf.d_tecn_prop_reg = row[24].value  # varchar2(10),
                    certProf.d_doc_req = row[25].value  # varchar2(100),
                    certProf.f_doc_req = row[26].value  # date,
                    certProf.f_doc_pres = row[27].value  # date,
                    certProf.f_propuesta = ""  # date,
                    certProf.f_res_informe = row[28].value  # date,
                    certProf.set_c_prop_aprob(row[29].value)  # X = True, ''= False
                    certProf.set_c_prop_denegada(row[30].value) #  varchar(50) X, DESISTIDO ART.71, DESISTIDO ART.91, DESISTIDO ART.68, DESISTIDO ART.94, ACUMULADO, SEPE,OTRA CA, Vacia (SE HA CAMBIADO EL TIPO EN BASE DE DATOS)
                    certProf.f_resol_deneg = row[31].value  # date,
                    certProf.f_notif_deneg = row[32].value  # date,
                    certProf.set_c_registro(row[34].value)  # varchar2(11),
                    certProf.f_resol_prov = row[35].value  # date,
                    certProf.f_recep_titulo = row[37].value  # date,
                    certProf.f_envio_titulo = row[38].value  # date,
                    certProf.f_entrega_titulo = row[39].value  # date,
                    certProf.f_recibi_titulo = row[40].value  # date
                    certProf.d_signatura_arch = row[41].value

                    certProf.set_expediente_comunes()
                    certProf.set_f_resolucion()
                    certProf.set_f_notificacion_boe()
                    certProf.set_f_desestimiento()
                    certProf.set_f_envio_otra_ca()
                    certProf.set_f_certificacion()
                    certProf.set_c_estado()
                    certProf.set_tipo_acreditacion()
                    certProf.set_c_res_informe()

                    certProf.normalize_dates()

                    if certProf.flagUnidadesFormativas:
                        certProf.set_unidades_formativas()

                    # Des/comentar Necesita para script SQL descomentar para produccion
                    print("DECLARE")
                    print("v_per_fo_cod NUMBER;")
                    print("v_sol_cod NUMBER;")
                    print("BEGIN")
                    print("BEGIN")
                    print("sql.....")
                    print("EXCEPTION WHEN no_data_found THEN")

                    #if not (certProf.existe_persona(cursorOracle)): # Comentar para producción
                    if True:                                         # Descomentar para producción
                        certProf.insertaPersona(conOracle, cursorOracle)
                        logger.info('Insertamos persona (' + certProf.d_dni + ')')
                    else:
                        logger.info('La persona ya existe')

                    # Des/comentar para produccion
                    print("END;")

                    if certProf.c_certificado:
                        if certProf.insertSiSolCertProf(conOracle, logger):
                            countOracle += 1
                            certProf.flagOtrosPresentada = False
                            certProf.set_documentacion_presentada(cursorOracle, conOracle)
                            certProf.flagOtrosRequerida = False
                            certProf.set_documentacion_requerida(cursorOracle, conOracle)
                            certProf.set_d_cod_certificado(row[23].value, cursorOracle, conOracle)
                            certProf.arrayDocPresentada.clear()
                        else:
                            pass
                    else:
                        if certProf.c_certificado is not None:
                            logger.error(fileName + " Fila: " + str(certProf.n_fila) + " La especialidad " + str(certProf.c_certificado) + " no existe en la BBDD")
                        else:
                            logger.error(
                                fileName + " Fila: " + str(certProf.n_fila) + " NO TIENE CODIGO DE CERTIFICADO ")

                    print("END;")
                    print("/")

    except ImportError:
        logger.error("File " + fileName + "cant be opened")
    finally:
        logger.info('Filas insertadas: ' + str(countOracle))
        logger.info('Filas recorridas: ' + str(idx))
        logger.info('Start Index: ' + str(START_INDEX))


# print("Valores no tratados en documentos")
# print(certProf.arrayOtros)
# print("Valores no tratados c_prop_deneg")
# print(certProf.arrayDenegada)
# print("Valores no tratados c_prop_aprob")
# print(certProf.arrayAprobada)
# print("Valores no identificados en BBDD de UC:")
# print(certProf.arrayUC)
# print("Valores no identificados en BBDD de Especialidades:")
# print(set(certProf.arrayEspecialidad))
# print("Valores no identificados en BBDD de Especialidades Practicas:")
# print(set(certProf.arrayEspecialidadPracticas))
# print("Valores no identificados en BBDD Localidades:")
# print(set(certProf.arrayLocalidadesUnicas))

logging.info('End scriptXLS.py \n')

# connection.close()



def imprimirColores():
    print("\n\n")
    print("Gris Oscuro: ")
    print("PatternType: " + ws['B85'].fill.patternType)
    print("fgColor: ")
    print(ws['B85'].fill.fgColor)
    print("bgColor: ")
    print(ws['B85'].fill.bgColor)
    print("Start Color: ")
    print(ws['B85'].fill.start_color)
    print("End Color: ")
    print(ws['B85'].fill.end_color)

    print("\n\n")
    print("Gris Claro: ")
    print("PatternType: ")
    print(ws['B84'].fill.patternType)
    print("fgColor: ")
    print(ws['B84'].fill.fgColor)
    print("bgColor: ")
    print(ws['B84'].fill.bgColor)
    print("Start Color: ")
    print(ws['B84'].fill.start_color)
    print("End Color: ")
    print(ws['B84'].fill.end_color)

